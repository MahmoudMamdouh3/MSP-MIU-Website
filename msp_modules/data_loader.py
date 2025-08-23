import os
import glob
import pandas as pd
from openpyxl import load_workbook

def read_excel_with_status(file_path, color_col_letter='D'):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    green = 'FF00FF00'
    red = 'FFFF0000'

    statuses = []
    rows = []

    for row in range(2, ws.max_row + 1):
        cell = ws[f'{color_col_letter}{row}']
        color = cell.fill.start_color.index

        if color == green:
            status = 'Accepted'
        elif color == red:
            status = 'Rejected'
        else:
            status = 'Review'

        statuses.append(status)
        row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        rows.append(row_data)

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    df = pd.DataFrame(rows, columns=headers)
    df['status'] = statuses
    return df

def read_csv_file(file_path):
    df = pd.read_csv(file_path)
    df['status'] = 'Review'
    return df

def drop_duplicate_columns(df):
    return df.loc[:, ~df.columns.duplicated()]

def load_and_merge_all_files(folder_path='data', color_col_letter='D'):
    folder_path = os.path.abspath(folder_path)
    all_files = glob.glob(os.path.join(folder_path, '*.*'))
    dfs = []
    original_order = None

    for idx, file_path in enumerate(all_files):
        ext = os.path.splitext(file_path)[1].lower()
        print(f"📄 قراءة الملف: {os.path.basename(file_path)}")
        try:
            if ext == '.xlsx':
                df = read_excel_with_status(file_path, color_col_letter=color_col_letter)
            elif ext == '.csv':
                df = read_csv_file(file_path)
            else:
                print(f"⚠️ تجاهل الملف ذو الامتداد غير المدعوم: {file_path}")
                continue

            df = drop_duplicate_columns(df)

            if idx == 0:
                original_order = df.columns.tolist()

            dfs.append(df)
        except Exception as e:
            print(f"❌ خطأ في قراءة الملف {file_path}: {e}")

    if dfs:
        merged_df = pd.concat(dfs, ignore_index=True, sort=True)
        if original_order:
            cols_ordered = [col for col in original_order if col in merged_df.columns]
            extra_cols = [col for col in merged_df.columns if col not in cols_ordered]
            merged_df = merged_df[cols_ordered + extra_cols]
        return merged_df
    else:
        return pd.DataFrame()

if __name__ == "__main__":
    merged_df = load_and_merge_all_files(folder_path='data', color_col_letter='D')
    if not merged_df.empty:
        output_file = 'data/merged_with_status.xlsx'
        merged_df.to_excel(output_file, index=False)
        print(f"✅ تم دمج الملفات وحفظ النتيجة في: {output_file}")
    else:
        print("❌ لم يتم دمج أي بيانات.")
